# Copyright 2019-present Infosys Limited    
# SPDX-License-Identifier: Apache-2.0   

import re
from template import Template
import os
from openpyxl import load_workbook

libData = {}
progData = {}
template = Template()
wb = load_workbook('dataModel/prototypeV8.xlsx')
class xlUtils:
    def __init__(self):
        pass
    
   
    def extractField(field):
        varName = ''
        varType = ''
        encodeCondition = ''
        decodeCondition = ''
        validation = ''
        count = ''
        mVarName = re.search('^(\w+)',str(field))
        mVarType = re.search(':Type\[(\w+)\]',str(field))
        mEnCond = re.search(':Condition\[([^\:]*)\]',str(field))
        mDeCond = re.search(':DecodeCondition\[([^\:]*)]',str(field))
        mValid = re.search(':Validation\[([^\:]+)\]',str(field))
        mCount = re.search(':Count\[([^\:]*)]',str(field))
        if mVarName:
            varName = mVarName.group(1)
          
        if mVarType:
            varType = mVarType.group(1)
           
        if mEnCond:
            encodeCondition = mEnCond.group(1)
            
        if mDeCond:
            decodeCondition = mDeCond.group(1)
        else:
            if mEnCond:
                decodeCondition = mEnCond.group(1)
            
            
        if mValid:
            validation = mValid.group(1)
            dataVarName = 'data.' + varName
            validation = re.sub(r'\$.',dataVarName,validation)

        if mCount:
            count =  mCount.group(1)
            
        return (varName,varType,encodeCondition,decodeCondition,validation,count)
    
    def lengthToType(byteLength):
        lTdict = {1:'Uint8',2:'Uint16',3:'Uint32',4:'Uint32',5:'Uint64',6:'Uint64',7:'Uint64',8:'Uint64',65535:'Uint8',9:'Uint8',12:'Uint8'}
        return lTdict[byteLength]
     
    def getByteLength(sheet,i): 
        length = 0
        lengthStr = ''
        cell_value_C = xlUtils.getCellValue(sheet,i,'C')
        octets =  str(cell_value_C) 
        
        mTo = re.search('to',str(octets)) 
        mVar = re.search('^variable\[(.*)]',str(octets))
        mToD = re.search('^(\d+) to (\d+)',str(octets))
        mToWD = re.search('^(\w) to (\w)\+(\d+)',str(octets))
        mToWWD = re.search('^(\w)\+(\d+) to (\w)\+(\d+)',str(octets))
        
        if not mTo:
            if mVar:
                lenStr = mVar.group(1)
                length = 65535
                lengthStr = lenStr
                                           
            elif xlUtils.getBitLength(sheet,i,'D') == 8:
                length = 1
                
        else:
            if mToD:
                toD = mToD.group(1)
                toD2 = mToD.group(2)
                length = int(toD2) - int(toD) + 1
                
            elif mToWD:
                ToWD = mToWD.group(3)
                length = int(ToWD) + 1
                            
            elif mToWWD:
                ToWWD2 = mToWWD.group(2)
                ToWWD4 = mToWWD.group(4)
                length = int(ToWWD4) - int(ToWWD2) + 1        
        return (length,lengthStr)
    
    def getBitLength(sheet,i,j):
        length = 0
        fieldName = xlUtils.getCellValue(sheet,i,j)
        mcolDtoK = re.search('[D-K]',j)
        if fieldName != None and mcolDtoK:
           for k in map(chr, range( ord(j), ord('L'))):
               fName = xlUtils.getCellValue(sheet,i,k)
              
               if length == 0:
                   length = length + 1
                   
               elif fName == None:
                   length = length + 1
            
               else:
                   break
        return length
    
    
    def getCellValue(sheet,row,column):
        colNum = 0
        chartoint = {'A' : 1,'B' : 2, 'C' : 3, 'D' : 4, 'E' : 5, 'F' : 6, 'G' : 7,
               'H' : 8, 'I' : 9, 'J' : 10, 'K' : 11, 'L' : 12, 'M' : 13, 'N' : 14,
               'O' : 15, 'P' : 16, 'Q' : 17, 'R' : 18, 'S' : 19, 'T' : 20, 'U' : 21,
               'V' : 22, 'W' : 23, 'X' : 24, 'Y' : 25, 'Z' : 26}
        
        colNum = chartoint[column]
        cellvalue = sheet.cell(row=row, column=colNum).value
        return cellvalue
    
    def getVarNameFromString(localString,typeName):
          
        localString = localString.lower()
        varName = ''
        localString = re.sub(r'\(.*\)'," ",localString)
        localString=localString.replace('/'," ")
        localString = re.sub('-'," ",localString)
        tokens = [x.capitalize() for x in localString.split(" ")]
        varName = varName.join(tokens)
        varName = varName[0].upper() + varName[1:]
        if typeName == 1:
            return varName
        else:
            return varName[0].lower() + varName[1:]
        
    def templateProcess(templateData,ttFileNameCpp,ttFileNameH,outputDirCpp,outputDirH):
        
        
        template = Template()
        
        if not os.path.exists(outputDirCpp):
            os.makedirs(outputDirCpp)
        outputFileNameCpp = templateData['fileName'] + '.cpp'
        
        template.__init__({'OUTPUT' : outputFileNameCpp, 'OUTPUT_PATH' : outputDirCpp})
        template.process(ttFileNameCpp, {'tempdata' : templateData})
        

        if not os.path.exists(outputDirH):
            os.makedirs(outputDirH)
        outputFileNameH = templateData['fileName'] + '.h'
        template.__init__({'OUTPUT' : outputFileNameH, 'OUTPUT_PATH' : outputDirH})
        template.process(ttFileNameH, {'tempdata' : templateData})
        
    def addToMakeSo(libName,objFile,srcFile):
       
        fileData = {}
        fileData['objFile'] = objFile
        fileData['sourceFile'] = srcFile
        if libName not in libData:
            lib = libName
            lib = re.sub(r'.so','',lib)
            libData[libName] = {}
            libData[libName]['libName'] = lib
            libData[libName]['fileList'] = []
        if fileData not in libData[libName]['fileList']:
            libData[libName]['fileList'].append(fileData)
       
    def addToMakeExe(progName,objFile,srcFile):
        fileData = {}
        fileData['objFile'] = objFile
        fileData['sourceFile'] = srcFile
        
        if progName not in progData:
            prog = progName
            prog = re.sub(r'.exe','',prog)
            progData[progName] = {}
            progData[progName]['progName'] = prog
            
        progData[progName]['fileList'] = []
        progData[progName]['fileList'].append(fileData)
        
    def addSoToMakeExe(progName,soName):
        if progName not in progData:
            prog = progName
            prog = re.sub(r'.exe','',prog)
            progData[progName] = {}
            progData[progName]['progName'] = prog
            
            progData[progName]['soList'] = []
        progData[progName]['soList'].append(soName)
       
    def generateMakeFile():
        makeFileData = {}
       
        for libr in libData.keys():
             makeFileData['libList'] = []
             makeFileData['libList'].append(libData[libr])
        for prog in progData.keys():
            makeFileData['progList']=[]
            makeFileData['progList'].append(progData[prog])
        
        
        ttFileName = 'tts/makefiletemplate.tt'
       
        fileName = 'Makefile'
        outputDir='../../src/gtpV2Codec'
        
        template.__init__({'OUTPUT' : fileName, 'OUTPUT_PATH' : outputDir})
        template.process(ttFileName, {'makefiledata': makeFileData})
        
xlUtils()
